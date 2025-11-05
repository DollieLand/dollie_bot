import re
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler, CallbackQueryHandler, ContextTypes, filters
)
import json
import threading
import time
import requests
import telegram.error
import logging
import asyncio

# ========== НАСТРОЙКИ ==========
TOKEN = "8259429897:AAGeYH9i-OTWiRqDChsOTBaLK18MP0g-APs"
ADMIN_ID = 8496308542

# ========== НАСТРОЙКИ ДЛЯ PYTHONANYWHERE ==========
BOT_API_URL = "http://DollieLand.pythonanywhere.com"
ADMIN_TOKEN = "dollie_secret_2024"

# ========== ХРАНИЛИЩА (в памяти) ==========
orders = {}             # {order_id: status_text}
order_requests = {}     # {order_id: user_id}
active_chats = {}       # {admin_id: user_id}
chat_links = {}         # {admin_id: user_id, user_id: admin_id}
tags = {}               # {order_id: tag_text}
admin_notes = {}        # {order_id: note_text}
user_messages = {}      # {user_id: [messages]}
subscribed_users = set()
user_names = {}
last_known_status = {}
order_dates = {}        # {order_id: datetime}
order_titles = {}       # {order_id: title_text}

# Флаг для блокировки одновременной синхронизации
sync_lock = threading.Lock()
last_sync_time = 0
SYNC_INTERVAL = 3  # секунды между синхронизациями

# ========== СТАТУСЫ ЗАКАЗОВ ==========
order_statuses = {
    "1": {"button": "В обработке", "message": "🌸 Ваш заказ зарегистрирован и находится в обработке. Мы скоро начнём его оформление."},
    "2": {"button": "Оформлен", "message": "✨ Отлично! Ваш заказ успешно оформлен и скоро будет отправлен."},
    "3": {"button": "Продавец отправил ваш заказ", "message": "📦 Продавец отправил ваш заказ. Следим за движением!"},
    "4": {"button": "Прибыл на склад в США и находится в обработке", "message": "🎉 Ура! Ваш заказ прибыл на наш склад в США и находится в обработке. Скоро подготовим его к отправке и рассчитаем итоговый вес."},
    "5": {"button": "Требуется оплата веса", "message": "💰 Для продолжения доставки необходимо оплатить вес заказа. Администратор свяжется с вами для уточнения деталей."},
    "6": {"button": "Вес оплачен", "message": "🌿 Оплата веса получена! Ваш заказ готовится к отправке со склада."},
    "7": {"button": "Заказ отправлен в РФ", "message": "✈️ Ваш заказ отправлен в Россию! Ориентировачное ожидание 4-5 недель."},
    "8": {"button": "Заказ прибыл в РФ", "message": "🇷🇺 Ваш заказ прибыл в Россию. Скоро он будет доставлен к нам."},
    "9": {"button": "Прибыл в Москву", "message": "🏙️ Отлично! Ваш заказ прибыл в Москву. Через 7-10 дней он будет передан посреднику."},
    "10": {"button": "Заказ у посредника", "message": "👤 Ваш заказ передан посреднику. Свяжитесь с администратором для уточнения адреса доставки."},
    "11": {"button": "Заказ отправлен посредником", "message": "🚚 Ваш заказ отправлен посредником и уже в пути к вам."},
    "12": {"button": "Заказ успешно получен", "message": "🎉 Поздравляем! Вы успешно получили свой заказ. Поделитесь впечатлениями!"}
}

# Группировка статусов для анализа
status_groups = {
    "Начальные": ["В обработке", "Оформлен"],
    "В пути": ["Продавец отправил ваш заказ", "Прибыл на склад в США и находится в обработке", "Заказ отправлен в РФ"],
    "Ожидание оплата": ["Требуется оплата веса"],
    "В России": ["Вес оплачен", "Заказ прибыл в РФ", "Прибыл в Москву"],
    "Финальные": ["Заказ у посредника", "Заказ отправлен посредником", "Заказ успешно получен"]
}

# ========== FAQ ==========
faq = {
    "Важно! 💡": (
        "🌸 Немного важной информации:\n\n"
        "Цены на международную логистику могут меняться довольно часто — "
        "поэтому указанные суммы являются ориентировочными. "
        "Если курс валют или тарифы перевозчиков немного изменятся, итоговая стоимость может быть скорректирована. "
        "Мы всегда стараемся уведомлять об этом заранее и сохранять для вас максимально выгодные условия 💕\n\n"
        "Спасибо за ваше доверие и терпение — оно очень помогает мне работать для вас с заботой 💜"
    ),
    "Доставка из США ✈️": (
        "💌 Доставка из США обычно занимает около 4–5 недель, но иногда бывают небольшие задержки — "
        "до 2 месяцев (это зависит от работы транспортных служб и таможни).\n\n"
        "Все новые куклы заказываются исключительно из США, чтобы гарантировать оригинальность и качество 🌷\n\n"
        "💗 Примерная стоимость веса:\n"
        "• Тяжелая кукла в коробке с упаковкой (~1 кг): около <b>1600 руб.</b>\n"
        "• Коробка с куклой полегче с упаковкой (~500 г): около <b>800 руб.</b>\n"
        "• Б/у кукла (без коробки): от <b>300 руб.</b>\n\n"
        "💡 <b>Важно!</b> Цены ориентировочные и зависят от текущего курса валют. Стоимость доставки коллекционных изданий (гробики, скуллекторы) и тяжелых посылок рассчитывается индивидуально.\n\n"
        "Я всегда стараюсь подобрать для вас самый надежный и выгодный вариант доставки!"
    ),
    "Доставка из Китая 📦": (
        "🎎 Из Китая можно заказать не только кукол, но и фигурки, аксессуары, редкие игрушки, одежду и даже милые наклейки!\n\n"
        "Мы сотрудничаем с Taobao, Poizon и других площадок, и помогаем безопасно оформить заказ 💗\n\n"
        "✨ Средний срок доставки составляет 3–4 недели, но возможны задержки из-за загруженности таможни.\n\n"
        "💗 Примерная стоимость веса:\n"
        " • Б/у кукла — от 500 руб.\n"
        " • 1 кг — примерно 1800 руб.\n\n"
        "Я стараюсь, чтобы каждая посылка пришла к вам быстро, бережно и в идеальном состоянии 🌸"
    ),
    "Бронь и предзаказы 🔐": (
        "🩷 Если вы хотите забронировать куклу из лота — пожалуйста, убедитесь, что ваше решение окончательное. "
        "Бронь фиксируется после согласования и действует 2 дня. "
        "Если за это время оплата не поступает и нет сообщения о задержке, бронь может быть автоматически снята.\n\n"
        "💖 Предзаказы оформляются с полной оплатой заранее.\n\n"
        "Спасибо за понимание🌷"
    ),
    "Политика возвратов 🔄": (
        "🌼 Политика возвратов при выкупе кукол:\n\n"
        "💗 Я всегда стараюсь, чтобы каждая кукла соответствовала фотографиям продавца. "
        "По вашему запросу можно получить дополнительные снимки или уточнения перед покупкой.\n\n"
        "✨ Возврат возможен только при уважительной причине и если кукла ещё находится на руках (не была отправлена). "
        "В таких случаях удерживается мой процент посредника.\n\n"
        "🌸 Пожалуйста, учитывайте:\n"
        " • Я не могу отвечать за работу почтовых служб, таможни или скрытые дефекты, которые не были указаны продавцом.\n"
        " • Решение о покупке остаётся за вами — если есть сомнения, лучше заранее уточнить детали.\n"
        " • Возврат по причинам вроде «передумал» или «ожидал другое» невозможен.\n\n"
        "💖 Благодарю за понимание и доверие💕"
    )
}

# ========== ТЕКСТЫ ДЛЯ ШАБЛОНОВ ==========
how_order_text = (
    "🌷 Сделать заказ совсем несложно!\n\n"
    "Просто напишите @Darielune — расскажите, какую куклу или товар ищете, и вместе подберём идеальный вариант 💕\n\n"
    "💖 Что нужно для начала:\n"
    " • ссылка на понравившуюся куклу или любой товар;\n"
    " • опишите, что вы ищете и в каком бюджете.\n\n"
    "Я выкупаю куклы с Amazon, eBay, а также с китайских площадок — Taobao, Poizon и других.\n\n"
    "🌸 Кроме того, можно выбрать куклу прямо из Telegram-канала — DollieLand Shop. "
)
order_process_text = (
    "💗 Процесс оформления заказа:\n\n"
    " • Напишите «бронь» под публикации с выбранной куклой или свяжитесь со мной через @Darielune.\n"
    " • После подтверждения будут отправлены реквизиты для оплаты.\n"
    " • Оплата за куклу производится сразу, а доставка по весу оплачивается после прибытия на склад.\n"
    " • Когда заказ приедет, он будет аккуратно упакован и отправлен удобным способом — СДЭК, Авито Доставка и др.\n\n"
    "🌷 Всё просто, прозрачно и с заботой — чтобы каждая кукла нашла свой дом 💕"
)
where_track_text = (
    "✨ После оформления заказа я отправлю вам трек-номер — с его помощью можно отслеживать посылку.\n\n"
    "💌 Если статус заказа ещё не определён — это значит, что заказ ещё не обработан в базе. Не волнуйтесь.\n\n"
    "🌿 Все обновления по заказу будут приходить автоматически в этот чат, чтобы вы всегда были в курсе о передвижениях своей посылки."
)

# ========== АВТООТВЕТЫ ==========
auto_replies = {
    "greeting": {
        "keywords": ["привет", "здравствуй", "здрасьте", "хай", "добрый день", "доброе утро", "добрый вечер", "здравствуйте"],
        "response": "Привет! 😊 Добро пожаловать в DollieBot!\nЯ могу помочь вам узнать статус заказа, стоимость доставки или ответить на любые вопросы по товарам. Напишите трек-номер или 'позови человека' если нужен оператор 💖"
    },
    "thanks": {
        "keywords": ["спасибо", "благодарю", "спс", "thx"],
        "response": "Всегда рада помочь! 💙 Если у вас возникнут новые вопросы по заказам, доставке или стоимости, не стесняйтесь писать."
    },
    "goodbye": {
        "keywords": ["пока", "до свидания", "увидимся", "bye", "до скорого"],
        "response": "До скорого! 👋 Буду рада помочь вам в любое время."
    },
    "delivery": {
        "keywords": ["доставка", "сроки", "доставить", "посылка", "отправка", "shipment", "tracking"],
        "response": "🌿 Информация о доставке: в среднем доставка занимает от 3 до 5 недель. Если вы оформляете заказ из США — срок обычно 4–5 недель, из Китая — около 3–4 недель."
    },
    "price": {
        "keywords": ["стоимость", "цена", "вес", "сумма", "расходы", "тариф"],
        "response": "💰 Стоимость заказа складывается из цены самого товара и доставки по весу. Примерно: 1 кг — 1600 руб (США), из Китая — ~1300 руб/кг."
    },
    "order": {
        "keywords": ["заказ", "оформить", "купить", "бронь", "предзаказ", "order", "booking"],
        "response": "📦 Чтобы сделать заказ, напишите @Darielune и опишите, какую куклу или товар вы ищете. Укажите ссылки, бюджет и пожелания."
    },
    "track": {
        "keywords": ["трек", "отследить", "tracking", "номер", "status"],
        "response": "📬 После оформления заказа вы получите трек-номер. Все обновления будут приходить автоматически в этот чат."
    }
}

# ========== УЛУЧШЕННЫЕ ФУНКЦИИ СИНХРОНИЗАЦИИ С ВЕБ-ПАНЕЛЬЮ ==========

def call_admin_api(endpoint, method="GET", data=None):
    """Вызов API веб-панели с улучшенной обработкой ошибок"""
    try:
        url = f"{BOT_API_URL}/{endpoint}"
        headers = {"X-Admin-Token": ADMIN_TOKEN}

        if method == "GET":
            response = requests.get(url, headers=headers, timeout=10)
        else:
            response = requests.post(url, json=data, headers=headers, timeout=10)

        if response.status_code == 200:
            return response.json()
        else:
            print(f"❌ API Error {response.status_code}: {response.text}")
            return {"ok": False, "error": f"HTTP {response.status_code}"}
    except Exception as e:
        print(f"🌐 API Connection Error: {e}")
        return {"ok": False, "error": str(e)}

def save_bot_data():
    """Сохраняет все данные бота в файл и синхронизирует с веб-панелью"""
    with sync_lock:
        try:
            data_to_save = {
                'orders': orders,
                'order_requests': order_requests,
                'admin_notes': admin_notes,
                'tags': tags,
                'user_names': user_names,
                'order_dates': {k: v.isoformat() for k, v in order_dates.items()},
                'last_known_status': last_known_status,
                'user_messages': user_messages,
                'active_chats': active_chats,
                'chat_links': chat_links,
                'order_titles': order_titles
            }

            # Сохраняем локально
            with open('bot_data.json', 'w', encoding='utf-8') as f:
                json.dump(data_to_save, f, ensure_ascii=False, indent=2)
            print(f"💾 Данные сохранены локально: {len(orders)} заказов, {len(order_titles)} названий")

            # Синхронизируем с веб-панелью
            sync_result = sync_with_web_panel()
            if sync_result:
                print("✅ Данные синхронизированы с веб-панелью")
            else:
                print("⚠️ Синхронизация с веб-панелью не удалась")

            return True
        except Exception as e:
            print(f"❌ Ошибка сохранения: {e}")
            return False

def sync_with_web_panel():
    """Синхронизирует данные с веб-панелью"""
    try:
        data_to_sync = {
            'orders': orders,
            'order_requests': order_requests,
            'admin_notes': admin_notes,
            'tags': tags,
            'user_names': user_names,
            'order_dates': {k: v.isoformat() for k, v in order_dates.items()},
            'order_titles': order_titles,
            'timestamp': datetime.now().isoformat()
        }

        result = call_admin_api("api/notify_update", method="POST", data=data_to_sync)

        if result.get('ok'):
            print(f"🔄 Данные отправлены в веб-панель: {len(orders)} заказов")
            return True
        else:
            print(f"❌ Ошибка синхронизации: {result.get('error', 'Unknown error')}")
            return False
    except Exception as e:
        print(f"❌ Ошибка при синхронизации: {e}")
        return False

def load_bot_data():
    """Загружает данные бота из веб-панели или локального файла"""
    global orders, order_requests, admin_notes, tags, user_names, order_dates
    global last_known_status, user_messages, active_chats, chat_links, order_titles

    with sync_lock:
        try:
            # Пробуем загрузить из веб-панели
            print("🔄 Загружаем данные из веб-панели...")
            api_result = call_admin_api("api/get_orders")

            if api_result and api_result.get('ok'):
                data = api_result.get('data', {})
                return load_data_from_dict(data, "веб-панели")

        except Exception as e:
            print(f"🌐 Ошибка подключения к веб-панели: {e}")

        # Если веб-панель недоступна, грузим из файла
        try:
            if os.path.exists('bot_data.json'):
                with open('bot_data.json', 'r', encoding='utf-8') as f:
                    data = json.load(f)
                return load_data_from_dict(data, "файла")
        except Exception as e:
            print(f"❌ Ошибка загрузки из файла: {e}")

        print("ℹ️ Начинаем с пустыми данными")
        return False

def load_data_from_dict(data, source):
    """Загружает данные из словаря в глобальные переменные"""
    global orders, order_requests, admin_notes, tags, user_names, order_dates
    global last_known_status, user_messages, active_chats, chat_links, order_titles

    # Очищаем и загружаем новые данные
    orders.clear()
    order_requests.clear()
    admin_notes.clear()
    tags.clear()
    user_names.clear()
    order_dates.clear()
    last_known_status.clear()
    user_messages.clear()
    active_chats.clear()
    chat_links.clear()
    order_titles.clear()

    orders.update(data.get('orders', {}))
    order_requests.update(data.get('order_requests', {}))
    admin_notes.update(data.get('admin_notes', {}))
    tags.update(data.get('tags', {}))
    user_names.update(data.get('user_names', {}))
    user_messages.update(data.get('user_messages', {}))
    order_titles.update(data.get('order_titles', {}))

    # Восстанавливаем даты
    for k, v in data.get('order_dates', {}).items():
        try:
            if isinstance(v, str):
                order_dates[k] = datetime.fromisoformat(v)
            else:
                order_dates[k] = v
        except:
            order_dates[k] = datetime.now()

    print(f"✅ Данные загружены из {source}: {len(orders)} заказов, {len(order_titles)} названий")
    return True

# Загружаем данные при запуске
load_bot_data()

# ========== ХЕЛПЕРЫ ==========
def save_user_message(user_id: int, text: str):
    lst = user_messages.setdefault(user_id, [])
    lst.append(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {text}")
    save_bot_data()

def build_status_keyboard(order_id: str):
    buttons = []
    items = list(order_statuses.items())
    row = []
    for i, (k, v) in enumerate(items, 1):
        row.append(InlineKeyboardButton(v["button"], callback_data=f"setstatus__{order_id}__{k}"))
        if i % 2 == 0:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    return InlineKeyboardMarkup(buttons)

def format_order_list_entry(order_id: str):
    status = orders.get(order_id, "не определён")
    owner_id = order_requests.get(order_id)

    if owner_id:
        username = user_names.get(owner_id)
        owner_repr = f"@{username} (ID:{owner_id})" if username else f"ID:{owner_id}"
    else:
        owner_repr = "—"

    tag = tags.get(order_id, "")
    tag_part = f" [{tag}]" if tag else ""

    title = order_titles.get(order_id, "")
    title_part = f" - {title}" if title else ""

    return f"{order_id}: {status} — {owner_repr}{tag_part}{title_part}"

def format_order_details(order_id: str, for_admin: bool = False):
    """Форматирует детальную информацию о заказе"""
    status = orders.get(order_id, "не определён")
    owner_id = order_requests.get(order_id)
    tag = tags.get(order_id, "")
    note = admin_notes.get(order_id, "")
    created_date = order_dates.get(order_id, datetime.now())
    title = order_titles.get(order_id, "")

    if owner_id:
        username = user_names.get(owner_id)
        owner_repr = f"@{username} (ID:{owner_id})" if username else f"ID:{owner_id}"
    else:
        owner_repr = "—"

    days_since_creation = (datetime.now() - created_date).days

    text = f"📦 Заказ {order_id}\n"

    if title:
        text += f"🏷️ Название: {title}\n"

    text += f"📊 Статус: {status}\n"
    text += f"👤 Владелец: {owner_repr}\n"
    text += f"📅 Создан: {created_date.strftime('%d.%m.%Y')} ({days_since_creation} дней назад)\n"

    if tag:
        text += f"🔖 Тег: {tag}\n"

    if for_admin and note:
        text += f"📝 Заметка: {note}\n"

    return text

# ========== КОМАНДЫ ДЛЯ НАЗВАНИЙ ЗАКАЗОВ ==========
async def add_title(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Добавить название для заказа (только для админа)"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав для этой команды.")
        return

    if len(context.args) < 2:
        await update.message.reply_text(
            "Использование: /title <order_id> <название заказа>\n"
            "Пример: /title 12345 Кукла Барби Одежда\n"
            "Пример: /title 12345 \"Комплект одежды для Барби\""
        )
        return

    order_id = context.args[0].strip()
    title_text = " ".join(context.args[1:]).strip()

    if order_id not in orders and order_id not in order_requests:
        await update.message.reply_text(f"❌ Заказ {order_id} не найден.")
        return

    order_titles[order_id] = title_text
    save_bot_data()

    await update.message.reply_text(f"✅ Название для заказа {order_id} добавлено:\n\"{title_text}\"")

async def clear_title(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Очистить название заказа (только для админа)"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав для этой команды.")
        return

    if not context.args:
        await update.message.reply_text("Использование: /cleartitle <order_id>")
        return

    order_id = context.args[0].strip()

    if order_id in order_titles:
        order_titles.pop(order_id)
        save_bot_data()
        await update.message.reply_text(f"✅ Название для заказа {order_id} очищено.")
    else:
        await update.message.reply_text(f"ℹ️ У заказа {order_id} нет названия.")

# ========== КОМАНДЫ ДЛЯ ЗАМЕТОК АДМИНА ==========
async def add_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Добавить заметку к заказу (только для админа)"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав для этой команды.")
        return

    if len(context.args) < 2:
        await update.message.reply_text(
            "Использование: /note <order_id> <текст заметки>\n"
            "Пример: /note 12345 Нужно уточнить адрес доставки"
        )
        return

    order_id = context.args[0].strip()
    note_text = " ".join(context.args[1:]).strip()

    if order_id not in orders and order_id not in order_requests:
        await update.message.reply_text(f"❌ Заказ {order_id} не найден.")
        return

    admin_notes[order_id] = note_text
    save_bot_data()
    await update.message.reply_text(f"✅ Заметка для заказа {order_id} добавлена:\n{note_text}")

async def view_notes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Просмотр всех заметок (только для админа)"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав для этой команды.")
        return

    if not admin_notes:
        await update.message.reply_text("📝 Заметок пока нет.")
        return

    text = "📝 Заметки к заказам:\n\n"
    for order_id, note in admin_notes.items():
        if note.strip():
            text += f"📦 {order_id}: {note}\n"

    if len(text) > 4000:
        for i in range(0, len(text), 4000):
            await update.message.reply_text(text[i:i+4000])
    else:
        await update.message.reply_text(text)

async def order_details(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Детальная информация о заказе"""
    if not context.args:
        await update.message.reply_text("Использование: /order <order_id>")
        return

    order_id = context.args[0].strip()

    if order_id not in orders and order_id not in order_requests:
        await update.message.reply_text(f"❌ Заказ {order_id} не найден.")
        return

    is_admin = (update.effective_user.id == ADMIN_ID)
    text = format_order_details(order_id, for_admin=is_admin)

    if is_admin:
        keyboard = [
            [
                InlineKeyboardButton("✏️ Изменить статус", callback_data=f"setstatus__{order_id}__menu"),
                InlineKeyboardButton("📝 Добавить заметку", callback_data=f"addnote__{order_id}")
            ],
            [
                InlineKeyboardButton("🏷️ Изменить тег", callback_data=f"edittag__{order_id}"),
                InlineKeyboardButton("📋 Добавить название", callback_data=f"addtitle__{order_id}")
            ],
            [
                InlineKeyboardButton("🗑️ Удалить заказ", callback_data=f"delete__{order_id}")
            ]
        ]
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard))
    else:
        await update.message.reply_text(text)

async def clear_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Очистить заметку к заказу (только для админа)"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав для этой команды.")
        return

    if not context.args:
        await update.message.reply_text("Использование: /clearnote <order_id>")
        return

    order_id = context.args[0].strip()

    if order_id in admin_notes:
        admin_notes.pop(order_id)
        save_bot_data()
        await update.message.reply_text(f"✅ Заметка для заказа {order_id} очищена.")
    else:
        await update.message.reply_text(f"ℹ️ У заказа {order_id} нет заметки.")

# ========== УЛУЧШЕННЫЙ ЭКСПОРТ В EXCEL ==========
async def export_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав для этой команды.")
        return

    # Принудительная синхронизация перед экспортом
    load_bot_data()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказы"

    # Стили
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    headers = [
        "ID заказа", "Название заказа", "Статус", "Группа статуса", "ID пользователя",
        "Username", "Тег", "Заметка админа", "Дата создания", "Дней с создания", "Приоритет"
    ]

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    all_order_ids = sorted(
        set(list(orders.keys()) + list(order_requests.keys())),
        key=lambda x: order_dates.get(x, datetime.min),
        reverse=True
    )

    for order_id in all_order_ids:
        status = orders.get(order_id, "не определён")
        user_id = order_requests.get(order_id)
        username = user_names.get(user_id, "")
        tag = tags.get(order_id, "")
        note = admin_notes.get(order_id, "")
        created_date = order_dates.get(order_id, datetime.now())
        title = order_titles.get(order_id, "")

        status_group = "Не определено"
        for group, statuses in status_groups.items():
            if status in statuses:
                status_group = group
                break

        days_since_creation = (datetime.now() - created_date).days

        if days_since_creation > 30 and status in ["В обработке", "Оформлен"]:
            priority = "🔴 Высокий"
        elif days_since_creation > 14 and status in status_groups["Начальные"]:
            priority = "🟡 Средний"
        else:
            priority = "🟢 Нормальный"

        ws.append([
            order_id, title, status, status_group, user_id or "", username,
            tag, note, created_date.strftime("%d.%m.%Y %H:%M"), days_since_creation, priority
        ])

    for row in range(2, len(all_order_ids) + 2):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col in [1, 4, 5, 9, 10, 11]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

            if col == 11:
                if cell.value == "🔴 Высокий":
                    cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                elif cell.value == "🟡 Средний":
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                elif cell.value == "🟢 Нормальный":
                    cell.fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

    column_widths = {
        "A": 12, "B": 25, "C": 25, "D": 15, "E": 12, "F": 15,
        "G": 15, "H": 30, "I": 16, "J": 12, "K": 12
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    table_ref = f"A1:K{len(all_order_ids) + 1}"
    table = Table(displayName="OrdersTable", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    ws.freeze_panes = "A2"

    filename = f"orders_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)

    try:
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=open(filename, "rb"),
            caption="📊 Полный отчет по заказам с улучшенной таблицей\n\n"
                   "📋 3 листа:\n"
                   "• Заказы - основная таблица с фильтрами\n"
                   "• Статистика - аналитика и метрики\n"
                   "• Анализ - заказы, требующие внимания"
        )
    finally:
        try:
            os.remove(filename)
        except Exception:
            pass

# ========== БЫСТРЫЙ ЭКСПОРТ ==========
async def export_quick(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав для этой команды.")
        return

    # Принудительная синхронизация перед экспортом
    load_bot_data()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказы"

    headers = ["ID заказа", "Название", "Статус", "Пользователь", "Тег", "Заметка", "Дата создания"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    all_order_ids = sorted(set(list(orders.keys()) + list(order_requests.keys())))
    for order_id in all_order_ids:
        status = orders.get(order_id, "не определён")
        user_id = order_requests.get(order_id)
        username = user_names.get(user_id, "") if user_id else ""
        tag = tags.get(order_id, "")
        note = admin_notes.get(order_id, "")
        created_date = order_dates.get(order_id, datetime.now())
        title = order_titles.get(order_id, "")

        user_info = f"@{username}" if username else f"ID:{user_id}" if user_id else ""

        ws.append([order_id, title, status, user_info, tag, note, created_date.strftime("%d.%m.%Y")])

    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 30)

    filename = f"orders_quick_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)

    try:
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=open(filename, "rb"),
            caption="📦 Быстрый экспорт заказов"
        )
    finally:
        try:
            os.remove(filename)
        except Exception:
            pass

async def unassigned_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показать заказы без привязанных пользователей (только для админа)"""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав для этой команды.")
        return

    # Принудительная синхронизация
    load_bot_data()

    unassigned = []
    for order_id, user_id in order_requests.items():
        if user_id is None:
            title = order_titles.get(order_id, "")
            status = orders.get(order_id, "не определён")
            unassigned.append((order_id, title, status))

    if not unassigned:
        await update.message.reply_text(
            "📦 <b>Заказы без привязанных пользователей</b>\n\n"
            "🌿 Все заказы имеют привязанных пользователей",
            parse_mode="HTML"
        )
        return

    text = "📦 <b>Заказы без привязанных пользователей:</b>\n\n"
    for order_id, title, status in unassigned[:10]:
        title_display = title if title else "❌ Без названия"
        text += f"🆔 <code>{order_id}</code>\n"
        text += f"🏷️ {title_display}\n"
        text += f"📊 {status}\n\n"

    if len(unassigned) > 10:
        text += f"💫 Показано 10 из {len(unassigned)} заказов\n\n"

    text += "💡 <i>Пользователи могут привязаться отправив ID заказа боту</i>"

    await update.message.reply_text(text, parse_mode="HTML")

# ========== ADMIN: создание заказа ==========
async def create_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав для этой команды.")
        return

    if not context.args:
        await update.message.reply_text("Использование: /create <order_id>")
        return

    order_id = context.args[0].strip()

    if order_id not in orders:
        orders[order_id] = "не определён"
        order_requests[order_id] = None
        order_dates[order_id] = datetime.now()
        save_bot_data()
        await update.message.reply_text(f"✅ Заказ {order_id} создан. Ожидается привязка пользователя.")
    else:
        await update.message.reply_text(f"ℹ️ Заказ {order_id} уже существует.")

# ========== ADMIN: удаление заказа ==========
async def delete_order_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав для этой команды.")
        return

    if not context.args:
        await update.message.reply_text("Использование: /delete <order_id>")
        return

    order_id = context.args[0].strip()

    if order_id not in orders and order_id not in order_requests:
        await update.message.reply_text(f"❌ Заказ {order_id} не найден.")
        return

    owner_id = order_requests.get(order_id)
    status = orders.get(order_id, "не определён")
    tag = tags.get(order_id, "")
    note = admin_notes.get(order_id, "")
    title = order_titles.get(order_id, "")

    orders.pop(order_id, None)
    order_requests.pop(order_id, None)
    tags.pop(order_id, None)
    admin_notes.pop(order_id, None)
    order_dates.pop(order_id, None)
    order_titles.pop(order_id, None)

    for user_id in list(last_known_status.keys()):
        if order_id in last_known_status[user_id]:
            last_known_status[user_id].pop(order_id)

    save_bot_data()

    response_text = f"✅ Заказ {order_id} удалён администратором.\n"
    response_text += f"📦 Статус был: {status}\n"
    if tag:
        response_text += f"🏷️ Тег был: {tag}\n"
    if note:
        response_text += f"📝 Заметка была: {note}\n"
    if title:
        response_text += f"📋 Название было: {title}\n"
    if owner_id:
        username = user_names.get(owner_id, "")
        response_text += f"👤 Владелец: {f'@{username}' if username else f'ID:{owner_id}'}"

        try:
            await context.bot.send_message(
                chat_id=owner_id,
                text=f"📦 Заказ {order_id} был удалён администратором. "
                     f"Если это ошибка, свяжитесь с администратором."
            )
        except Exception:
            pass

    await update.message.reply_text(response_text)

# ========== USER: удаление своего заказа ==========
async def delete_order_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    if not context.args:
        await update.message.reply_text("Использование: /mydelete <order_id>")
        return

    order_id = context.args[0].strip()

    if order_id not in order_requests:
        await update.message.reply_text(f"❌ Заказ {order_id} не найден.")
        return

    if order_requests.get(order_id) != user_id:
        await update.message.reply_text(
            f"❌ Заказ {order_id} не принадлежит вам. "
            f"Вы можете удалять только свои заказы."
        )
        return

    current_status = orders.get(order_id, "не определён")
    non_deletable_statuses = [
        "Продавец отправил ваш заказ",
        "Прибыл на склад в США и находится в обработке",
        "Требуется оплата веса",
        "Вес оплачен",
        "Заказ отправлен в РФ",
        "Заказ прибыл в РФ",
        "Прибыл в Москву",
        "Заказ у посредника",
        "Заказ отправлен посредником",
        "Заказ успешно получен"
    ]

    if any(status in current_status for status in non_deletable_statuses):
        await update.message.reply_text(
            f"❌ Невозможно удалить заказ {order_id}.\n"
            f"📦 Текущий статус: {current_status}\n\n"
            f"Заказ уже находится в процессе доставки. "
            f"Для отмены заказа свяжитесь с администратором."
        )
        return

    status = orders.get(order_id, "не определён")
    tag = tags.get(order_id, "")
    title = order_titles.get(order_id, "")

    orders.pop(order_id, None)
    order_requests.pop(order_id, None)
    tags.pop(order_id, None)
    order_dates.pop(order_id, None)
    order_titles.pop(order_id, None)

    if user_id in last_known_status and order_id in last_known_status[user_id]:
        last_known_status[user_id].pop(order_id)

    save_bot_data()

    username = user_names.get(user_id, "")
    try:
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"👤 Пользователь удалил свой заказ!\n"
                 f"Пользователь: @{username or 'None'} (ID:{user_id})\n"
                 f"Заказ: {order_id}\n"
                 f"Статус был: {status}"
        )
    except Exception:
        pass

    await update.message.reply_text(
        f"✅ Ваш заказ {order_id} удалён.\n"
        f"📦 Статус был: {status}\n\n"
        f"Если это ошибка, свяжитесь с администратором."
    )

# ========== ADMIN: массовое удаление заказов ==========
async def delete_orders_bulk(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав для этой команды.")
        return

    if not context.args:
        await update.message.reply_text(
            "Использование: /deletebulk <order_id1> <order_id2> ...\n"
            "Пример: /deletebulk 12345 67890 11121"
        )
        return

    deleted_count = 0
    error_count = 0
    results = []

    for order_id in context.args:
        order_id = order_id.strip()

        if order_id not in orders and order_id not in order_requests:
            results.append(f"❌ {order_id}: не найден")
            error_count += 1
            continue

        owner_id = order_requests.get(order_id)
        status = orders.get(order_id, "не определён")
        title = order_titles.get(order_id, "")

        orders.pop(order_id, None)
        order_requests.pop(order_id, None)
        tags.pop(order_id, None)
        admin_notes.pop(order_id, None)
        order_dates.pop(order_id, None)
        order_titles.pop(order_id, None)

        for user_id in list(last_known_status.keys()):
            if order_id in last_known_status[user_id]:
                last_known_status[user_id].pop(order_id)

        results.append(f"✅ {order_id}: удалён (статус: {status})")
        deleted_count += 1

        if owner_id:
            try:
                await context.bot.send_message(
                    chat_id=owner_id,
                    text=f"📦 Заказ {order_id} был удалён администратором. "
                         f"Если это ошибка, свяжитесь с администратором."
                )
            except Exception:
                pass

    save_bot_data()

    report = f"📊 Результат массового удаления:\n"
    report += f"✅ Удалено: {deleted_count}\n"
    report += f"❌ Ошибок: {error_count}\n\n"
    report += "\n".join(results)

    await update.message.reply_text(report)

# ========== ОБРАБОТЧИКИ КОМАНД ==========
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    welcome = (
        "🎀 <b>Добро пожаловать в DollieHelper!</b> 🎀\n\n"
        "✨ Здесь вы можете:\n"
        "— отслеживать свои заказы,\n"
        "— узнать всё о доставке и оплате,\n"
        "— или просто задать вопрос — и мы обязательно вам ответим 💕\n\n"
        "📦 <b>Хотите узнать статус заказа?</b>\n"
        "Отправьте трек-номер отдельным сообщением — и бот подскажет, где сейчас ваша посылочка 🎁\n\n"
        "💬 <b>Нужна помощь живого человека?</b>\n"
        "Просто напишите «позови человека» или нажмите кнопку ниже — и скоро с вами свяжутся 💫\n\n"
        "Ниже — несколько быстрых подсказок 🌿"
    )
    buttons = [
        [InlineKeyboardButton("Написать @Darielune💌 ", url="https://t.me/Darielune")],
        [InlineKeyboardButton("Как оформить заказ 💕", callback_data="how_order")],
        [InlineKeyboardButton("Как получить трек-номер 📬", callback_data="where_track")],
        [InlineKeyboardButton("FAQ — ответы 🌿", callback_data="show_faq")],
        [InlineKeyboardButton("Мои заказы 📦", callback_data="my_orders")],
        [InlineKeyboardButton("Позвать человека 💬", callback_data="call_admin")]
    ]
    await update.message.reply_text(welcome, parse_mode="HTML", reply_markup=InlineKeyboardMarkup(buttons))

async def faq_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [[InlineKeyboardButton(k, callback_data=f"faq__{k}")] for k in faq.keys()]
    await update.message.reply_text("💡 Часто задаваемые вопросы — просто выберите интересующий пункт 🌸", reply_markup=InlineKeyboardMarkup(keyboard))

async def my_orders_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    # Принудительная синхронизация перед показом заказов
    load_bot_data()

    user_orders = [oid for oid, uid in order_requests.items() if uid == user_id]

    if not user_orders:
        await update.message.reply_text(
            "🌿 Похоже, у вас пока нет активных заказов.\nНо ничего страшного — если что-то понравится, можно оформить заказ в любое время 💕"
        )
        return

    text = "📦 <b>Ваши заказы:</b>\n\n"
    user_statuses = last_known_status.setdefault(user_id, {})

    for order_id in user_orders:
        current_status = orders.get(order_id, "не определён")
        previous_status = user_statuses.get(order_id)
        title = order_titles.get(order_id, "")

        text += f"🆔 <b>Заказ #{order_id}</b>\n"
        if title:
            text += f"🏷️ <b>{title}</b>\n"

        if previous_status and previous_status != current_status:
            text += f"📊 Статус: <s>{previous_status}</s> → <b>{current_status}</b>\n\n"
        else:
            text += f"📊 Статус: <b>{current_status}</b>\n\n"

        user_statuses[order_id] = current_status

    text += "💫 <i>Для подробной информации по конкретному заказу используйте:</i>\n"
    text += "<code>/order номер_заказа</code>\n\n"
    text += "🔔 <i>Статусы обновляются автоматически. Следите за уведомлениями!</i>"

    await update.message.reply_text(text, parse_mode="HTML")

# ========== ADMIN: orders/status/broadcast/etc. ==========
async def orders_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав для этой команды.")
        return

    # Принудительная синхронизация
    load_bot_data()

    args = context.args
    results = []
    try:
        if len(args) >= 2 and args[0].lower() == "user":
            uid = int(args[1])
            for oid, u in order_requests.items():
                if u == uid:
                    results.append(format_order_list_entry(oid))
        elif args:
            status_filter = " ".join(args).lower()
            for oid, st in orders.items():
                if st.lower() == status_filter:
                    results.append(format_order_list_entry(oid))
        else:
            seen = set(list(orders.keys()) + list(order_requests.keys()))
            for oid in sorted(seen):
                results.append(format_order_list_entry(oid))
    except Exception as ex:
        await update.message.reply_text(f"Ошибка при разборе аргументов: {ex}")
        return
    if not results:
        await update.message.reply_text("📦 Заказы не найдены по фильтру.")
    else:
        text = "\n".join(results)
        chunk_size = 3000
        for i in range(0, len(text), chunk_size):
            await update.message.reply_text(text[i:i+chunk_size])

async def broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав.")
        return
    message = " ".join(context.args)
    if not message:
        await update.message.reply_text("Использование: /broadcast <сообщение>")
        return
    sent = 0
    unique_users = set(u for u in order_requests.values() if u)
    for uid in unique_users:
        try:
            await context.bot.send_message(chat_id=uid, text=f"📢 Сообщение от администрации:\n\n{message}")
            sent += 1
        except Exception:
            pass
    await update.message.reply_text(f"✅ Отправлено: {sent} пользователям.")

async def active_chats_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав.")
        return
    if not active_chats:
        await update.message.reply_text("🌿 Сейчас нет активных диалогов.")
        return
    text = "🟢 Активные диалоги:\n"
    for admin, uid in active_chats.items():
        text += f"Админ {admin} ↔ Пользователь {uid}\n"
    await update.message.reply_text(text)

async def exportchat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав.")
        return
    try:
        uid = int(context.args[0])
    except Exception:
        await update.message.reply_text("Использование: /exportchat <user_id>")
        return
    msgs = user_messages.get(uid, [])
    if not msgs:
        await update.message.reply_text("Нет сообщений с этим пользователем.")
        return
    filename = f"chat_{uid}_{datetime.now().strftime('%Y%m%d%H%M%S')}.txt"
    with open(filename, "w", encoding="utf-8") as f:
        f.write("\n".join(msgs))
    try:
        await context.bot.send_document(chat_id=update.effective_chat.id, document=open(filename, "rb"))
    finally:
        try:
            os.remove(filename)
        except Exception:
            pass

# ========== ADMIN: update status ==========
async def stats_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав.")
        return

    # Принудительная синхронизация
    load_bot_data()

    total_orders = len(orders)
    active_users = len(set(u for u in order_requests.values() if u))
    orders_with_notes = len([note for note in admin_notes.values() if note.strip()])
    orders_with_titles = len([title for title in order_titles.values() if title.strip()])
    status_count = {}
    for s in orders.values():
        status_count[s] = status_count.get(s, 0) + 1
    text = f"📊 Статистика:\nВсего заказов: {total_orders}\nАктивных пользователей: {active_users}\nЗаказов с заметками: {orders_with_notes}\nЗаказов с названиями: {orders_with_titles}\n\nСтатусы:\n"
    for s, cnt in status_count.items():
        text += f"{s}: {cnt}\n"
    await update.message.reply_text(text)

async def tag_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав.")
        return
    try:
        order_id = context.args[0]
        tag_text = " ".join(context.args[1:])
        if not tag_text:
            raise ValueError
    except Exception:
        await update.message.reply_text("Использование: /tag <order_id> <тег>")
        return
    tags[order_id] = tag_text
    save_bot_data()
    await update.message.reply_text(f"✅ Заказ {order_id} отмечен тегом: {tag_text}")

async def update_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав.")
        return

    order_id = None
    new_status = None

    if context.args:
        order_id = context.args[0]
        if len(context.args) >= 2:
            new_status = " ".join(context.args[1:]).strip()

    if (not order_id or not order_id.isdigit()) and update.message.reply_to_message:
        replied_text = update.message.reply_to_message.text.strip()
        if replied_text.isdigit():
            order_id = replied_text
            if len(context.args) >= 2:
                new_status = " ".join(context.args[1:]).strip()

    if not order_id or not new_status:
        raw = update.message.text or ""
        raw = re.sub(r'^/update(@\w+)?\s*', '', raw, count=1).strip()
        m = re.match(r'^(\d+)\s+(.+)$', raw, flags=re.DOTALL)
        if m:
            order_id = order_id or m.group(1)
            new_status = new_status or m.group(2).strip()

    if not order_id:
        await update.message.reply_text("Использование: /update <order_id> <new_status> — укажите ID заказа.")
        return
    if not new_status:
        await update.message.reply_text("Использование: /update <order_id> <new_status> — укажите новый статус (может содержать пробелы).")
        return

    orders[order_id] = new_status
    save_bot_data()
    owner = order_requests.get(order_id)
    notify_text = f"📬 Статус заказ {order_id} обновлён: {new_status}"
    if owner:
        try:
            await context.bot.send_message(chat_id=owner, text=notify_text)
        except Exception as e:
            await update.message.reply_text(f"❌ Не удалось уведомить пользователя: {e}")

    await update.message.reply_text(f"✅ {notify_text} (отправлено владельцу: {owner})")

# ========== ADMIN: обновление статуса через кнопки ==========
async def setstatus_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав.")
        return

    if context.args:
        order_id = context.args[0].strip()
    elif update.message.reply_to_message and update.message.reply_to_message.text and update.message.reply_to_message.text.strip().isdigit():
        order_id = update.message.reply_to_message.text.strip()
    else:
        await update.message.reply_text("Использование: /setstatus <order_id> — либо укажите ID в аргументе, либо ответьте на сообщение с ID заказа.")
        return

    if order_id not in orders:
        orders[order_id] = "не определён"
    if order_id not in order_requests:
        order_requests[order_id] = None
    if order_id not in order_dates:
        order_dates[order_id] = datetime.now()

    await update.message.reply_text(f"Выберите новый статус для заказа {order_id}:", reply_markup=build_status_keyboard(order_id))

# ========== CALLBACKS ==========
async def callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Полный обработчик callback-запросов от inline кнопок"""
    query = update.callback_query
    await query.answer()
    data = query.data or ""

    print(f"🔍 Callback data: {data}")

    try:
        if data == "my_orders":
            await my_orders_cmd(query, context)
            return

        elif data == "show_faq":
            keyboard = [[InlineKeyboardButton(k, callback_data=f"faq__{k}")] for k in faq.keys()]
            keyboard.append([InlineKeyboardButton("📂 В главное меню", callback_data="main_menu")])
            await query.edit_message_text(
                "💡 Часто задаваемые вопросы — просто выберите интересующий пункт 🌸",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return

        elif data.startswith("faq__"):
            faq_key = data.split("__", 1)[1]
            if faq_key in faq:
                await query.edit_message_text(
                    faq[faq_key],
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup([
                        [InlineKeyboardButton("📂 Назад к FAQ", callback_data="show_faq")],
                        [InlineKeyboardButton("📂 В главное меню", callback_data="main_menu")]
                    ])
                )
            else:
                await query.answer("❌ Вопрос не найден", show_alert=True)
            return

        elif data == "how_order":
            await query.edit_message_text(
                how_order_text,
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("📬 Получить трек-номер", callback_data="where_track")],
                    [InlineKeyboardButton("💌 Написать @Darielune", url="https://t.me/Darielune")],
                    [InlineKeyboardButton("📂 В главное меню", callback_data="main_menu")]
                ])
            )
            return

        elif data == "order_process":
            await query.edit_message_text(
                order_process_text,
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("💌 Написать @Darielune", url="https://t.me/Darielune")],
                    [InlineKeyboardButton("📂 Назад", callback_data="how_order")],
                    [InlineKeyboardButton("📂 В главное меню", callback_data="main_menu")]
                ])
            )
            return

        elif data == "where_track":
            await query.edit_message_text(
                where_track_text,
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("💌 Написать @Darielune", url="https://t.me/Darielune")],
                    [InlineKeyboardButton("📂 Назад", callback_data="how_order")],
                    [InlineKeyboardButton("📂 В главное меню", callback_data="main_menu")]
                ])
            )
            return

        elif data == "call_admin":
            user_id = query.from_user.id
            username = query.from_user.username or f"id{user_id}"

            active_chats[ADMIN_ID] = user_id
            chat_links[user_id] = ADMIN_ID
            chat_links[ADMIN_ID] = user_id

            try:
                await context.bot.send_message(
                    chat_id=ADMIN_ID,
                    text=f"👤 Пользователь @{username or user_id} (ID:{user_id}) просит подключиться к чату через кнопку."
                )
            except Exception:
                pass

            await query.edit_message_text(
                "🌸 Администратор будет уведомлён. Всё, что вы напишете далее, будет автоматически пересылаться администратору.",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("📂 В главное меню", callback_data="main_menu")]
                ])
            )
            return

        elif data == "main_menu":
            welcome = (
                "🎀 <b>Добро пожаловать в DollieHelper!</b> 🎀\n\n"
                "✨ Здесь вы можете:\n"
                "— отслеживать свои заказы,\n"
                "— узнать всё о доставке и оплате,\n"
                "— или просто задать вопрос — и мы обязательно вам ответим 💕\n\n"
                "📦 <b>Хотите узнать статус заказа?</b>\n"
                "Отправьте трек-номер отдельным сообщением — и бот подскажет, где сейчас ваша посылочка 🎁\n\n"
                "💬 <b>Нужна помощь живого человека?</b>\n"
                "Просто напишите «позови человека» или нажмите кнопку ниже — и с вами скоро свяжутся 💫\n\n"
                "Ниже — несколько быстрых подсказок 🌿"
            )
            buttons = [
                [InlineKeyboardButton("Написать @Darielune💌 ", url="https://t.me/Darielune")],
                [InlineKeyboardButton("Как оформить заказ 💕", callback_data="how_order")],
                [InlineKeyboardButton("Как получить трек-номер 📬", callback_data="where_track")],
                [InlineKeyboardButton("FAQ — ответы 🌿", callback_data="show_faq")],
                [InlineKeyboardButton("Мои заказы 📦", callback_data="my_orders")],
                [InlineKeyboardButton("Позвать человека 💬", callback_data="call_admin")]
            ]
            await query.edit_message_text(welcome, parse_mode="HTML", reply_markup=InlineKeyboardMarkup(buttons))
            return

        await query.answer("❌ Команда не распознана", show_alert=True)

    except Exception as e:
        print(f"❌ Ошибка в callback_handler: {e}")
        await query.answer("❌ Произошла ошибка при обработке запроса", show_alert=True)

# ========== ADMIN: завершение чата ==========
async def end_chat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав.")
        return

    if ADMIN_ID in active_chats:
        uid = active_chats.pop(ADMIN_ID)
        chat_links.pop(uid, None)
        chat_links.pop(ADMIN_ID, None)
        try:
            await context.bot.send_message(chat_id=uid, text="🌸 Администратор завершил чат.")
        except Exception:
            pass
        await update.message.reply_text(f"✅ Чат с пользователем ID:{uid} завершён.")
    else:
        await update.message.reply_text("🌿 Нет активного чата.")

# ========== ADMIN: команда reply ==========
async def reply_to_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Нет прав.")
        return
    if not context.args:
        await update.message.reply_text("Использование: /reply <user_id> <текст>")
        return
    try:
        target = int(context.args[0])
    except Exception:
        await update.message.reply_text("Неверный user_id.")
        return
    text = " ".join(context.args[1:]).strip()
    if not text:
        await update.message.reply_text("Укажите текст сообщения.")
        return
    try:
        await context.bot.send_message(chat_id=target, text=f"👤 Администратор: {text}")
        await update.message.reply_text("✅ Отправлено.")

        active_chats[ADMIN_ID] = target
        chat_links[target] = ADMIN_ID
        chat_links[ADMIN_ID] = target

        await update.message.reply_text(
            f"🌿 Все последующие сообщения администратора будут автоматически пересылаться пользователю (ID:{target}) "
            f"до завершения диалога командой /end."
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка при отправке: {e}")

# ========== УЛУЧШЕННАЯ АВТОСИНХРОНИЗАЦИЯ ==========
def auto_sync_data():
    """Фоновая синхронизация с веб-панелью"""
    global last_sync_time

    while True:
        time.sleep(SYNC_INTERVAL)

        with sync_lock:
            try:
                current_time = time.time()
                if current_time - last_sync_time < SYNC_INTERVAL:
                    continue

                api_result = call_admin_api("api/get_orders")
                if api_result.get('ok'):
                    data = api_result.get('data', {})

                    web_orders = data.get('orders', {})
                    web_titles = data.get('order_titles', {})
                    web_order_requests = data.get('order_requests', {})

                    # Проверяем реальные изменения
                    has_changes = (
                        web_orders != orders or
                        web_titles != order_titles or
                        web_order_requests != order_requests
                    )

                    if has_changes:
                        print("🔄 Обнаружены изменения в данных, синхронизируем...")
                        load_data_from_dict(data, "веб-панели (автосинхронизация)")
                        last_sync_time = current_time

            except Exception as e:
                print(f"⚠️ Ошибка автосинхронизации: {e}")

# Запускаем фоновую синхронизацию
sync_thread = threading.Thread(target=auto_sync_data, daemon=True)
sync_thread.start()

# ========== ОБРАБОТЧИК СООБЩЕНИЙ ==========
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик текстовых сообщений от пользователей"""

    user_id = update.effective_user.id
    text = update.message.text.strip()
    username = update.effective_user.username or f"id{user_id}"

    print(f"🔍 Получено сообщение от {user_id}: '{text}'")

    save_user_message(user_id, text)

    if user_id in chat_links:
        target = chat_links[user_id]
        prefix = ""
        if user_id == ADMIN_ID:
            prefix = "👤 Администратор:"
        else:
            prefix = f"💬 Пользователь @{username or user_id}:"
        try:
            await context.bot.send_message(chat_id=target, text=f"{prefix}\n{text}")
        except Exception:
            pass
        return

    text_lower = text.lower()

    for reply_key, reply_data in auto_replies.items():
        for kw in reply_data["keywords"]:
            if kw.lower() in text_lower:
                print(f"✅ Найдено ключевое слово: '{kw}', отвечаю автоответом")
                await update.message.reply_text(reply_data["response"])
                return

    if text.isdigit() or text.startswith('DOLL'):
        order_id = text
        current_owner = order_requests.get(order_id)

        if current_owner is None:
            order_requests[order_id] = user_id
            user_names[user_id] = username
            subscribed_users.add(user_id)
            is_new_attachment = True

            if order_id not in order_dates:
                order_dates[order_id] = datetime.now()

            print(f"🔗 Пользователь {user_id} (@{username}) привязан к заказу {order_id}")

        elif current_owner == user_id:
            is_new_attachment = False
        else:
            await update.message.reply_text(
                f"❌ Заказ {order_id} уже привязан к другому пользователю. "
                f"Обратитесь к администратору для уточнения деталей."
            )
            return

        if order_id not in orders:
            orders[order_id] = "не определён"

        if is_new_attachment:
            try:
                title = order_titles.get(order_id, "")
                title_info = f"\n🏷️ Название: {title}" if title else ""

                await context.bot.send_message(
                    chat_id=ADMIN_ID,
                    text=f"🔗 <b>Пользователь привязан к заказу!</b>\n\n"
                         f"👤 Пользователь: @{username or 'None'} (ID:{user_id})\n"
                         f"📦 Заказ: {order_id}{title_info}\n"
                         f"📊 Статус: {orders.get(order_id, 'не определён')}",
                    parse_mode="HTML"
                )
            except Exception as e:
                print(f"❌ Ошибка уведомления админа: {e}")

        current_status = orders.get(order_id, "не определён")
        user_statuses = last_known_status.setdefault(user_id, {})
        previous_status = user_statuses.get(order_id)

        title = order_titles.get(order_id, "")
        title_message = f"\n🏷️ <b>{title}</b>" if title else ""

        if is_new_attachment:
            await update.message.reply_text(
                f"✅ <b>Вы успешно привязаны к заказу {order_id}!</b>{title_message}\n"
                f"📦 Текущий статус: {current_status}\n\n"
                f"💫 Теперь вы будете получать уведомления об изменении статуса",
                parse_mode="HTML"
            )
        elif previous_status and previous_status != current_status:
            await update.message.reply_text(
                f"📬 <b>Ваш заказ {order_id}</b>{title_message}\n"
                f"🔄 Статус изменен: '{previous_status}' → '{current_status}'",
                parse_mode="HTML"
            )
        else:
            await update.message.reply_text(
                f"📦 <b>Заказ {order_id}</b>{title_message}\n"
                f"📊 Статус: {current_status}",
                parse_mode="HTML"
            )

        user_statuses[order_id] = current_status
        save_bot_data()
        return

    admin_keywords = ["позови человек", "позови человека", "админ", "оператор", "помощь", "свяжи", "человек"]
    if any(keyword in text_lower for keyword in admin_keywords):
        active_chats[ADMIN_ID] = user_id
        chat_links[user_id] = ADMIN_ID
        chat_links[ADMIN_ID] = user_id
        try:
            await context.bot.send_message(
                chat_id=ADMIN_ID,
                text=f"👤 Пользователь @{username or user_id} (ID:{user_id}) просит подключиться к чату.\n\nСообщение: {text}"
            )
        except Exception:
            pass
        await update.message.reply_text(
            "🌸 Администратор будет уведомлён. Всё, что вы напишете далее, будет автоматически пересылаться администратору."
        )
        return

    keyboard = [
        [InlineKeyboardButton("💌 Написать @Darielune", url="https://t.me/Darielune")],
        [InlineKeyboardButton("❓ Как заказать", callback_data="how_order")],
        [InlineKeyboardButton("📦 Мои заказы", callback_data="my_orders")],
        [InlineKeyboardButton("📂 В главное меню", callback_data="main_menu")]
    ]

    await update.message.reply_text(
        "✨ Пожалуйста, введите трек-номер для отслеживания посылки. 📦\n\n"
        "Или напишите 'позови человека' для связи с админом.",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

# ========== ОБРАБОТЧИК ОШИБОК ==========
async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик ошибок бота"""
    try:
        error_msg = f"⚠️ Ошибка бота: {context.error}"
        print(error_msg)

        if hasattr(context.error, '__class__'):
            print(f"🔧 Тип ошибки: {context.error.__class__.__name__}")

        if isinstance(context.error, telegram.error.NetworkError):
            print("🌐 Проблема с сетью, продолжаем работу...")
            return

        elif isinstance(context.error, telegram.error.TimedOut):
            print("⏰ Таймаут запроса, продолжаем работу...")
            return

        elif isinstance(context.error, telegram.error.Conflict):
            print("🔄 Конфликт - другой экземпляр бота уже запущен")
            return

        elif isinstance(context.error, telegram.error.BadRequest):
            print(f"❌ BadRequest: {context.error}")
            return

        if update and update.effective_message:
            try:
                await update.effective_message.reply_text(
                    "❌ Произошла временная ошибка. Пожалуйста, попробуйте еще раз через несколько секунд."
                )
            except:
                pass

    except Exception as e:
        print(f"💥 Ошибка в обработчике ошибок: {e}")

async def my_orders_cmd(query, context):
    """Обработчик кнопки Мои заказы через callback"""
    user_id = query.from_user.id

    # Принудительная синхронизация
    load_bot_data()

    user_orders = [oid for oid, uid in order_requests.items() if uid == user_id]

    if not user_orders:
        await query.edit_message_text(
            "🌿 Похоже, у вас пока нет активных заказов.\nНо ничего страшного — если что-то понравится, можно оформить заказ в любое время 💕",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("💌 Написать @Darielune", url="https://t.me/Darielune")],
                [InlineKeyboardButton("📂 В главное меню", callback_data="main_menu")]
            ])
        )
        return

    text = "📦 <b>Ваши заказы:</b>\n\n"
    user_statuses = last_known_status.setdefault(user_id, {})

    for order_id in user_orders:
        current_status = orders.get(order_id, "не определён")
        previous_status = user_statuses.get(order_id)
        title = order_titles.get(order_id, "")

        text += f"🆔 <b>Заказ #{order_id}</b>\n"
        if title:
            text += f"🏷️ <b>{title}</b>\n"

        if previous_status and previous_status != current_status:
            text += f"📊 Статус: <s>{previous_status}</s> → <b>{current_status}</b>\n\n"
        else:
            text += f"📊 Статус: <b>{current_status}</b>\n\n"

        user_statuses[order_id] = current_status

    text += "💫 <i>Для подробной информации по конкретному заказу используйте:</i>\n"
    text += "<code>/order номер_заказа</code>\n\n"
    text += "🔔 <i>Статусы обновляются автоматически. Следите за уведомлениями!</i>"

    await query.edit_message_text(
        text,
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("📂 В главное меню", callback_data="main_menu")],
            [InlineKeyboardButton("🔄 Обновить", callback_data="my_orders")]
        ])
    )

# ========== MAIN ==========
def main():
    print("🤖 Запускаем бота с улучшенной синхронизацией...")

    load_bot_data()

    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(MessageHandler(
        filters.TEXT & ~filters.COMMAND,
        handle_message
    ))

    # пользовательские команды
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("faq", faq_command))
    app.add_handler(CommandHandler("myorders", my_orders_cmd))
    app.add_handler(CommandHandler("mydelete", delete_order_user))
    app.add_handler(CommandHandler("order", order_details))

    # админские команды
    app.add_handler(CommandHandler("create", create_order))
    app.add_handler(CommandHandler("orders", orders_status))
    app.add_handler(CommandHandler("broadcast", broadcast))
    app.add_handler(CommandHandler("activechats", active_chats_cmd))
    app.add_handler(CommandHandler("exportchat", exportchat))
    app.add_handler(CommandHandler("stats", stats_cmd))
    app.add_handler(CommandHandler("tag", tag_order))
    app.add_handler(CommandHandler("update", update_status))
    app.add_handler(CommandHandler("setstatus", setstatus_cmd))
    app.add_handler(CommandHandler("reply", reply_to_user))
    app.add_handler(CommandHandler("end", end_chat))
    app.add_handler(CommandHandler("export_orders", export_orders))
    app.add_handler(CommandHandler("export_quick", export_quick))
    app.add_handler(CommandHandler("delete", delete_order_admin))
    app.add_handler(CommandHandler("deletebulk", delete_orders_bulk))

    # команды для заметок
    app.add_handler(CommandHandler("note", add_note))
    app.add_handler(CommandHandler("notes", view_notes))
    app.add_handler(CommandHandler("clearnote", clear_note))

    # команды для названий
    app.add_handler(CommandHandler("title", add_title))
    app.add_handler(CommandHandler("cleartitle", clear_title))

    # команда для незакрепленных заказов
    app.add_handler(CommandHandler("unassigned", unassigned_orders))

    app.add_error_handler(error_handler)

    app.add_handler(CallbackQueryHandler(callback_handler))

    print("✅ Бот запущен с улучшенной синхронизацией!")

    max_retries = 5
    retry_count = 0

    while retry_count < max_retries:
        try:
            app.run_polling(
                poll_interval=3,
                timeout=20,
                drop_pending_updates=True
            )
        except telegram.error.NetworkError as e:
            retry_count += 1
            print(f"🌐 Сетевая ошибка ({retry_count}/{max_retries}): {e}")
            if retry_count < max_retries:
                print("🔄 Перезапуск через 10 секунд...")
                time.sleep(10)
            else:
                print("❌ Достигнут лимит попыток перезапуска")
                break
        except Exception as e:
            print(f"❌ Критическая ошибка: {e}")
            print("🔄 Перезапуск через 10 секунд...")
            time.sleep(10)

if __name__ == "__main__":
    main()